import os
import csv
from datetime import datetime
from fastapi import FastAPI
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel
import uvicorn

from openpyxl import Workbook
from openpyxl.styles import PatternFill

app = FastAPI()

# =========================
# MODEL
# =========================
class PLCData(BaseModel):
    temp: float
    hum: float
    dew: float


latest_data = {}
history = []

ALARM_TEMP = 30


# =========================
# CSV ZAPIS
# =========================
def save_to_csv(record):
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"data_{today}.csv"
    file_exists = os.path.isfile(filename)

    with open(filename, mode="a", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file, delimiter=';')

        if not file_exists:
            writer.writerow([
                "Data i czas",
                "Temperatura [°C]",
                "Wilgotność [%]",
                "Punkt rosy [°C]"
            ])

        full_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        temp = str(round(record["temp"], 2)).replace(".", ",")
        hum = str(round(record["hum"], 2)).replace(".", ",")
        dew = str(round(record["dew"], 2)).replace(".", ",")

        writer.writerow([
            full_datetime,
            temp,
            hum,
            dew
        ])


# =========================
# GENEROWANIE EXCEL
# =========================
def generate_excel_report():
    today = datetime.now().strftime("%Y-%m-%d")
    csv_filename = f"data_{today}.csv"
    excel_filename = f"report_{today}.xlsx"

    if not os.path.isfile(csv_filename):
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Raport"

    red_fill = PatternFill(
        start_color="FFFF9999",
        end_color="FFFF9999",
        fill_type="solid"
    )

    with open(csv_filename, newline='', encoding="utf-8-sig") as file:
        reader = csv.reader(file, delimiter=';')

        for row_index, row in enumerate(reader, start=1):
            for col_index, value in enumerate(row, start=1):
                cell = ws.cell(row=row_index, column=col_index)
                cell.value = value

                # Kolumna 2 = temperatura
                if row_index > 1 and col_index == 2:
                    try:
                        temp_value = float(str(value).replace(",", "."))
                        if temp_value > ALARM_TEMP:
                            cell.fill = red_fill
                    except:
                        pass

    wb.save(excel_filename)
    return excel_filename


# =========================
# ODBIÓR DANYCH
# =========================
@app.post("/api/data")
async def receive_data(data: PLCData):
    global latest_data, history

    now = datetime.now()

    record = {
        "time": now.strftime("%H:%M:%S"),
        "timestamp_raw": now.timestamp(),
        "temp": round(data.temp, 2),
        "hum": round(data.hum, 2),
        "dew": round(data.dew, 2)
    }

    latest_data = record
    history.append(record)

    if len(history) > 100:
        history.pop(0)

    save_to_csv(record)

    return {"status": "ok"}


# =========================
# HISTORIA
# =========================
@app.get("/api/history")
async def get_history():
    return history


# =========================
# DOWNLOAD
# =========================
@app.get("/download")
async def download_file():
    excel_file = generate_excel_report()

    if excel_file and os.path.isfile(excel_file):
        return FileResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=excel_file
        )

    return {"error": "Brak danych do raportu"}


# =========================
# DASHBOARD
# =========================
@app.get("/", response_class=HTMLResponse)
async def dashboard():
    return """
<!DOCTYPE html>
<html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Monitoring PLC</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>

<style>
body { background:#0d1117; color:white; font-family:Arial; margin:0; }
h1 { text-align:center; padding:20px; }
.top-bar { text-align:center; margin-bottom:20px; }
.top-bar a { color:cyan; text-decoration:none; }
.values { display:flex; justify-content:center; gap:30px; flex-wrap:wrap; margin-bottom:30px; }
.box { background:#161b22; padding:18px 30px; border-radius:10px; font-size:24px; transition:0.3s; }
.alarm { background:#5c0000 !important; box-shadow:0 0 20px red !important; }
.chart-wrapper { max-width:1000px; margin:30px auto; }
canvas { width:100%!important; height:250px!important; }
</style>
</head>

<body>

<h1>📡 MONITORING PLC</h1>

<div class="top-bar">
<a href="/download">⬇ Pobierz raport Excel</a>
</div>

<div class="values">
<div class="box" id="tempBox">🌡 Temp<br><span id="tempNow">--</span> °C</div>
<div class="box">💧 Wilg<br><span id="humNow">--</span> %</div>
<div class="box">🌫 Rosa<br><span id="dewNow">--</span> °C</div>
</div>

<div class="chart-wrapper"><canvas id="tempChart"></canvas></div>
<div class="chart-wrapper"><canvas id="humChart"></canvas></div>
<div class="chart-wrapper"><canvas id="dewChart"></canvas></div>

<script>
function createChart(id,label,color){
    return new Chart(document.getElementById(id),{
        type:'line',
        data:{labels:[],datasets:[{label:label,data:[],borderColor:color,tension:0.4,pointRadius:0}]},
        options:{responsive:true,maintainAspectRatio:false,animation:false}
    });
}

const tempChart=createChart("tempChart","Temperatura (°C)","lime");
const humChart=createChart("humChart","Wilgotność (%)","cyan");
const dewChart=createChart("dewChart","Punkt rosy (°C)","orange");

async function update(){
    const res=await fetch('/api/history');
    const data=await res.json();
    if(data.length===0)return;

    const labels=data.map(d=>d.time);
    const temps=data.map(d=>d.temp);
    const hums=data.map(d=>d.hum);
    const dews=data.map(d=>d.dew);

    const last=data[data.length-1];

    document.getElementById("tempNow").innerText=last.temp.toFixed(2);
    document.getElementById("humNow").innerText=last.hum.toFixed(2);
    document.getElementById("dewNow").innerText=last.dew.toFixed(2);

    const tempBox=document.getElementById("tempBox");
    if(last.temp>30){tempBox.classList.add("alarm");}
    else{tempBox.classList.remove("alarm");}

    function apply(chart,values){
        chart.data.labels=labels;
        chart.data.datasets[0].data=values;
        chart.update();
    }

    apply(tempChart,temps);
    apply(humChart,hums);
    apply(dewChart,dews);
}

setInterval(update,2000);
update();
</script>

</body>
</html>
"""

# =========================
# START
# =========================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    uvicorn.run("app:app", host="0.0.0.0", port=port)
